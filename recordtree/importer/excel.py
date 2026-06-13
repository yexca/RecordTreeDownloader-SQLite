from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path

from openpyxl import load_workbook

from recordtree.exceptions import ImportRowError, ValidationError
from recordtree.models import ImportRecord
from recordtree.normalizers import clean_text, normalize_date
from recordtree.sizes import parse_size_text

from .parsers import parse_mega_json


HEADER_ALIASES = {
    "actor_raw": {"actor_raw", "actor", "演员", "声优", "澹颁紭"},
    "delivery_date": {"delivery_date", "配信日期", "閰嶄俊鏃ユ湡"},
    "title": {"title", "标题", "鏍囬"},
    "entry_date": {"entry_date", "录入日期", "褰曞叆鏃ユ湡"},
    "note": {"note", "备注", "澶囨敞"},
    "upload_title": {"upload_title", "上传标题", "涓婁紶鏍囬"},
    "duplicate_search_raw": {"duplicate_search_raw", "重复检索", "閲嶅妫€绱"},
    "source_name": {"source_name", "来源", "鏉ユ簮"},
    "mega_json": {"mega_json", "MEGA", "mega"},
    "size_raw": {"size_raw", "容量", "瀹归噺"},
}

REQUIRED_FIELDS = {
    "actor_raw",
    "delivery_date",
    "title",
    "entry_date",
    "note",
    "upload_title",
    "duplicate_search_raw",
    "source_name",
    "mega_json",
    "size_raw",
}


class ExcelImporter:
    def __init__(self) -> None:
        self.extra_columns: tuple[str, ...] = ()

    def count_records(self, path: Path) -> int:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            if sheet.max_row is None:
                return 0
            return max(sheet.max_row - 1, 0)
        finally:
            workbook.close()

    def iter_records(self, path: Path) -> Iterator[ImportRecord | ImportRowError]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            try:
                header_row = next(rows)
            except StopIteration as error:
                raise ValidationError("Workbook has no header row.") from error
            mapping, self.extra_columns = _build_header_mapping(header_row)
            for row_number, values in enumerate(rows, start=2):
                if _row_is_empty(values):
                    continue
                try:
                    yield _parse_row(values, mapping, row_number)
                except ImportRowError as error:
                    yield error
        finally:
            workbook.close()


def _build_header_mapping(header_row: tuple[object, ...]) -> tuple[dict[str, int], tuple[str, ...]]:
    seen: set[str] = set()
    mapping: dict[str, int] = {}
    extras: list[str] = []
    for index, raw_header in enumerate(header_row):
        header = clean_text(raw_header)
        if header is None:
            continue
        if header in seen:
            raise ValidationError(f"Duplicate Excel header: {header}")
        seen.add(header)
        field = _field_for_header(header)
        if field is None:
            extras.append(header)
            continue
        if field in mapping:
            raise ValidationError(f"Duplicate Excel header for field {field}: {header}")
        mapping[field] = index

    missing = sorted(REQUIRED_FIELDS - mapping.keys())
    if missing:
        raise ValidationError(f"Missing required Excel columns: {', '.join(missing)}")
    return mapping, tuple(extras)


def _field_for_header(header: str) -> str | None:
    normalized = header.strip().casefold()
    for field, aliases in HEADER_ALIASES.items():
        if normalized in {alias.casefold() for alias in aliases}:
            return field
    return None


def _parse_row(values: tuple[object, ...], mapping: dict[str, int], row_number: int) -> ImportRecord:
    def value(field: str) -> object | None:
        index = mapping[field]
        return values[index] if index < len(values) else None

    required_values = {
        "actor_raw": value("actor_raw"),
        "title": value("title"),
        "upload_title": value("upload_title"),
        "source_name": value("source_name"),
        "mega_json": value("mega_json"),
    }
    for field, raw in required_values.items():
        if clean_text(raw) is None:
            raise ImportRowError(
                "required_field_missing",
                f"Required field is missing: {field}",
                row_number=row_number,
                raw_value=field,
            )

    try:
        delivery_date = normalize_date(value("delivery_date"))
        entry_date = normalize_date(value("entry_date"))
    except ValidationError as error:
        raise ImportRowError(
            "date_invalid",
            str(error),
            row_number=row_number,
            raw_value=value("delivery_date") or value("entry_date"),
        ) from error

    size_raw = clean_text(value("size_raw"))
    mega_payload = parse_mega_json(value("mega_json"), row_number)
    return ImportRecord(
        source_type="xlsx",
        actor_raw=clean_text(value("actor_raw")) or "",
        delivery_date=delivery_date,
        title=clean_text(value("title")) or "",
        entry_date=entry_date,
        note=clean_text(value("note")),
        upload_title=clean_text(value("upload_title")) or "",
        duplicate_search_raw=clean_text(value("duplicate_search_raw")),
        source_name=clean_text(value("source_name")) or "",
        size_raw=size_raw,
        size_bytes=parse_size_text(size_raw),
        mega_file_name=mega_payload.file_names,
        mega_total_bytes=mega_payload.total_bytes,
        mega_formatted_size=mega_payload.formatted_size,
        mega_json=clean_text(value("mega_json")),
        source_row_number=row_number,
        links=mega_payload.links,
    )


def _row_is_empty(values: tuple[object, ...]) -> bool:
    return all(clean_text(value) is None for value in values)
